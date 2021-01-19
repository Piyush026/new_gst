import openpyxl
import requests

path = "Company and TAN.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
# print(m_row)
lst = []
for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    lst.append(cell_obj.value)
# print(lst)

# hu = [lst[i:i + 30] for i in range(0, len(lst), 30)]
# print(len(hu))

# url = "http://proxy.link/list/get/e40abb7f2ead053db343dd8e82e04ad7?geo=true"
# response = requests.get(url)
# proxy_list = (response.text.split('\n'))
# print(proxy_list)
# print(len(proxy_list))
# ji = zip(lst, proxy_list)
# print(ji)
# for x in ji:
#     print(x)

