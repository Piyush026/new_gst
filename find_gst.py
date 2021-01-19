# import random
# import time
#
# from openpyxl import load_workbook
# import pandas as pd
# import openpyxl
# import getpass
# import requests
# from openpyxl import Workbook
# from selenium.webdriver.common.keys import Keys
#
# from hhub import proxylist
#
# path = "CompanyInfo.xlsx"
# from config import (
#     get_web_driver_options,
#     get_chrome_web_driver,
#     set_browser_as_incognito,
#     set_ignore_certificate_error,
#     set_browser_in_fullScreen,
#     # set_automation_as_head_less,
#     URL, URL1
# )
# from selenium.webdriver.common.proxy import Proxy, ProxyType
# from selenium import webdriver
#
# dictt = []
#
#
# def fileread(gst_no):
#     wb_obj = openpyxl.load_workbook(path)
#     sheet_obj = wb_obj.active
#     m_row = sheet_obj.max_row
#     # print(m_row)
#     lst = []
#     for i in range(2, m_row + 1):
#         cell_obj = sheet_obj.cell(row=i, column=4)
#         lst.append(cell_obj.value)
#         sheet_obj.cell(row=i, column=3).value = gst_no
#     wb_obj.save(path)
#     # print(lst)
#     return lst
#
#
# # fileread()
#
#
# class GSTFINDER:
#
#     def __init__(self, base_url, pproxy):
#         self.base_url = base_url
#         prox = Proxy()
#         prox.proxy_type = ProxyType.MANUAL
#         prox.http_proxy = pproxy
#         # prox.socks_proxy = pproxy
#         prox.ssl_proxy = pproxy
#         print(pproxy)
#         capabilities = webdriver.DesiredCapabilities.CHROME
#         prox.add_to_capabilities(capabilities)
#
#         options = get_web_driver_options()
#         # set_proxy(options)
#         set_ignore_certificate_error(options)
#
#         set_browser_as_incognito(options)
#         # set_automation_as_head_less(options)
#         set_browser_in_fullScreen(options)
#
#         self.driver = get_chrome_web_driver(options, capabilities)
#
#     def know_your_gst(self, cname):
#         self.driver.get(self.base_url)
#         self.driver.implicitly_wait(5)
#         time.sleep(1)
#         self.driver.find_element_by_id("gstnumber").send_keys(cname + Keys.ENTER)
#         gst = self.driver.find_element_by_xpath('//*[@id="searchresult"]/span/strong[2]').text
#         dictt.append(gst)
#         # print(gst)
#         return gst
#
#     def close_driver(self):
#         self.driver.close()
#
#
# if __name__ == '__main__':
#     wb_obj = openpyxl.load_workbook(path)
#     sheet_obj = wb_obj.active
#     m_row = sheet_obj.max_row
#     # print(m_row)
#     lst = []
#     for idx in range(2, m_row + 1):
#         try:
#             cell_obj = sheet_obj.cell(row=idx, column=4)
#             # print(cell_obj.value)
#             proxy_list = proxylist()
#             pproxy = random.choice(proxy_list)
#             obj = GSTFINDER(URL, pproxy)
#             gst = obj.know_your_gst(cell_obj.value)
#             obj.close_driver()
#             sheet_obj.append(row=idx, column=3).value = gst
#             print(gst)
#
#         except:
#             continue
#         finally:
#             wb_obj.save(path)
#     wb_obj.save(path)

import openpyxl
import requests

path = "CompanyInfo.xlsx"

wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
# print(m_row)
lst = []
for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=4)
    lst.append(cell_obj.value)
print(len(lst))
print(lst)

# lst1 = [634,5467,3645657]
# hu = {lst: lst1}
# print(hu)