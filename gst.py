import csv
import json
import random
import threading
import time

import openpyxl
from selenium.webdriver.common.keys import Keys
from sql_connection import insert_row
import save_db
from hhub import proxylist

path = "CompanyInfo.xlsx"
from config import (
    get_web_driver_options,
    get_chrome_web_driver,
    set_browser_as_incognito,
    set_ignore_certificate_error,
    set_browser_in_fullScreen, URL,
    set_automation_as_head_less,
)
from selenium.webdriver.common.proxy import Proxy, ProxyType
from selenium import webdriver


def readfile():
    wb_obj = openpyxl.load_workbook(path)

    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    # print(m_row)
    lst = []
    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=4)
        lst.append(cell_obj.value)
    # print(len(lst))
    # print(lst)
    return lst


dictt = []
threadLocal = threading.local()


class GSTFINDER:

    def __init__(self, base_url, pproxy):
        self.base_url = base_url
        prox = Proxy()
        prox.proxy_type = ProxyType.MANUAL
        prox.http_proxy = pproxy
        # prox.socks_proxy = pproxy
        prox.ssl_proxy = pproxy
        print(pproxy)
        self.driver = getattr(threadLocal, 'driver', None)
        if self.driver is None:
            capabilities = webdriver.DesiredCapabilities.CHROME
            prox.add_to_capabilities(capabilities)

            options = get_web_driver_options()
            # set_proxy(options)
            set_ignore_certificate_error(options)

            set_browser_as_incognito(options)
            set_automation_as_head_less(options)
            # set_browser_in_fullScreen(options)

            self.driver = get_chrome_web_driver(options, capabilities)
        setattr(threadLocal, 'driver', self.driver)

    def know_your_gst(self, cname):
        try:
            # self.driver.set_page_load_timeout(10)
            self.driver.get(self.base_url)
            self.driver.implicitly_wait(5)
            time.sleep(1)
            self.driver.find_element_by_id("gstnumber").send_keys(cname + Keys.ENTER)
            gst = self.driver.find_element_by_xpath('//*[@id="searchresult"]/span/strong[2]').text
            dictt.append(gst)
            # print(gst)
            return gst
        except Exception as e:
            print(str(e))

    def close_driver(self):
        self.driver.close()


def maain(lst):
    cmpdict = {}
    try:
        proxy_list = proxylist()
        pproxy = random.choice(proxy_list)
        obj = GSTFINDER(URL, pproxy)
        gst = obj.know_your_gst(lst[1])
        print(gst)
        # obj.close_driver()
        cmpdict["company"] = lst[1]
        cmpdict["gst"] = gst
        # cmpdict[x] = gst
        print(cmpdict)
        """add for sql"""
        data = (lst[0], lst[1], gst)
    except:
        obj.close_driver()
        # continue

    finally:
        if cmpdict["gst"]:
            # csv_columns = ["company", "GST"]
            # csv_file = "cmp_gst.csv"
            # with open(csv_file, 'a') as csvfile:
            #     w = csv.DictWriter(csvfile, cmpdict.keys())
            #     w.writerow(cmpdict)
            # save_db.insertData(cmpdict)
            insert_row(data)


if __name__ == '__main__':
    # lsttt = ["ASHIRVAD PIPES PRIVATE LIMITED ", "ASCO NUMATICS (INDIA) PRIVATE LIMITED ", "sdjzhcb",
    #          "ARMAN FINANCIAL SERVICES LIMITED"]
    lsttt = [['U00339AN1986PLC007658' 'HIGHTECH PISTON (KARNATAKA) LIMITED   '],
             ['U67120AN1994PTC000043' 'ANDAMAN OVERSEAS FINVEST PRIVATE LIMITED   '],
             ['U65923AN2015PTC000260'
              'ISLANDERS MARGIN FREE FINSERVICE PRIVATE LIMITED  ']]
    maain(lsttt)
